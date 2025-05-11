import flet as ft

import openpyxl

import sqlite3
  
  
def main(page: ft.Page):
    page.title = "Multi-Page App"
    dbname = 'option.db'
    conn = sqlite3.connect(dbname)
    cur = conn.cursor()
    setting = None


    def navigate_to(page_name):
        global setting
        cur.execute('SELECT * FROM setting')
        setting = cur.fetchall()

        if page_name == "home":
            page.views.clear()
            page.views.append(home_view())
        elif page_name == "page1":
            page.views.clear()
            page.views.append(page1_view())
        elif page_name == "page2":
            page.views.clear()
            page.views.append(page2_view())
        elif page_name == "page3":
            page.views.clear()
            page.views.append(page3_view())
        elif page_name == "setting":
            page.views.clear()
            page.views.append(setting_view())
        page.update()

    def home_view():
        SpecificationPath = ft.TextField(
            label="Path",
            value="",
        ) 
        SCSpath = ft.TextField(
            label="Path",
            value="",
        )
        return ft.View(
            "/",
            appbar=ft.AppBar(
                title=ft.Text("Home"),
                actions=[
                    ft.IconButton(ft.Icons.HOME, on_click=lambda _: navigate_to("home")),
                    ft.IconButton(ft.Icons.FIBER_NEW, on_click=lambda _: navigate_to("page1")),
                    ft.IconButton(ft.Icons.EDIT_NOTE, on_click=lambda _: navigate_to("page2")),
                    ft.IconButton(ft.Icons.LOOKS_3, on_click=lambda _: navigate_to("page3")),
                    ft.IconButton(ft.Icons.SETTINGS_OUTLINED, on_click=lambda _: navigate_to("setting")),
                ],
            ),
            controls=[
                ft.Text("Welcome to the Home Page!", size=30),
                ft.Row(
                    controls=[
                        ft.Column(
                            controls=[
                                ft.Text("製作明細表", size=30),
                                SpecificationPath,
                            ],
                            expand=True
                        ),
                        ft.Column(
                            controls=[
                                ft.Text("System SCS", size=30),
                                SCSpath,
                            ],
                            expand=True
                        ),
                    ],
                    scroll=ft.ScrollMode.ALWAYS,
                    expand=True
                ),
            ],
        )
    
    def page1_view():
        def read_option(e):
            for i in range(setting[0][0]):
                chip = ft.Chip(
                    label=ft.Text("Option" + str(i+1)),
                    selected_color= ft.Colors.AMBER,
                    data = i+1,
                    selected = False,
                    on_click=lambda e: on_option(e),
                    width=300,
                )
                cells = [ft.DataCell(chip), ft.DataCell(ft.Text(value="Option" + str(i),))]
                rows.append(ft.DataRow(cells=cells))
                
        def on_option(e):
            pass
        global setting
        option = ft.TextField(
            label="条件",
            value="",
        )
        ms = ft.RadioGroup(
            content=ft.Column(
                controls=[
                    ft.Radio(value="1", label="メカオプション"),
                    ft.Radio(value="2", label="ソフトオプション"),
                ],
            ),
        )
        rows = []
        header = [ft.DataColumn(ft.Text("ON/OFF")), ft.DataColumn(ft.Text("OptionName"))]
        data_table = ft.DataTable(columns=header, rows=rows)
        
        return ft.View(
            "/page1",
            appbar=ft.AppBar(
                title=ft.Text("オプション対応編集"),
                actions=[
                    ft.IconButton(ft.Icons.HOME, on_click=lambda _: navigate_to("home")),
                    ft.IconButton(ft.Icons.FIBER_NEW, on_click=lambda _: navigate_to("page1")),
                    ft.IconButton(ft.Icons.EDIT_NOTE, on_click=lambda _: navigate_to("page2")),
                    ft.IconButton(ft.Icons.LOOKS_3, on_click=lambda _: navigate_to("page3")),
                    ft.IconButton(ft.Icons.SETTINGS_OUTLINED, on_click=lambda _: navigate_to("setting")),
                ],
            ),
            controls=[
                ft.Row(
                    controls=[
                        ft.Column(
                            controls=[
                                option,
                                ms,
                                ft.ElevatedButton("Read", on_click=lambda e: print("Read")),
                                ft.ElevatedButton("Save", on_click=lambda e: print("Save")),
                            ],
                            expand=True
                        ),
                    ],
                    scroll=ft.ScrollMode.ALWAYS,
                    expand=True
                ),
            ],
        )

    def page2_view():
        def on_option(e):
            pass

        def read_data(e):
            opcount = 1
            rows = []
            # for data in alldata:
            #     for i in range(16):
            #         chip = ft.Chip(
            #             label=ft.Text("Option" + str(opcount)),
            #             selected_color= ft.Colors.AMBER,
            #             data = opcount,
            #             selected = (data[i] == "1"),
            #             on_click=lambda e: on_option(e),
            #             width=300,
            #         )
            #         opcount += 1
            #         cells = [ft.DataCell(chip), ft.DataCell(ft.Text(value="Option" + str(i),))]
            #         rows.append(ft.DataRow(cells=cells))
            data_table.rows = rows
            page.update()
        

        header = [ft.DataColumn(ft.Text("ON/OFF")), ft.DataColumn(ft.Text("OptionName"))]
        rows = []
        data_table = ft.DataTable(columns=header, rows=rows)

        editfile_PATH = ft.TextField(
            label="パス",
            value="",
        )
        
        ReadButton = ft.Button("読み込み", on_click = lambda e : read_data(e))
        savebotton = ft.Button("保存", on_click = lambda e : print("保存"))

        return ft.View(
            "/page2",
            appbar=ft.AppBar(
                title=ft.Text("Page 2"),
                actions=[
                    ft.IconButton(ft.Icons.HOME, on_click=lambda _: navigate_to("home")),
                    ft.IconButton(ft.Icons.FIBER_NEW, on_click=lambda _: navigate_to("page1")),
                    ft.IconButton(ft.Icons.EDIT_NOTE, on_click=lambda _: navigate_to("page2")),
                    ft.IconButton(ft.Icons.LOOKS_3, on_click=lambda _: navigate_to("page3")),
                    ft.IconButton(ft.Icons.SETTINGS_OUTLINED, on_click=lambda _: navigate_to("setting")),
                ],
            ),
            controls=[
                ft.Text("This is Page 2", size=30),
                ft.Row(
                    controls=[
                        ft.Column(
                            controls=[data_table],
                            scroll=ft.ScrollMode.ALWAYS,
                            expand=True
                        ),
                    ],
                    scroll=ft.ScrollMode.ALWAYS,
                    expand=True
                ),
            ],
        )

    def page3_view():
        Path = ft.TextField(
            label="Path",
            value="",
        )
        def read_path(e):
            path = Path.value
            if path[0] == "\"" and path[-1] == "\"":
                path = path[1:-1]
            wb = openpyxl.load_workbook(path)
            sheet = wb[wb.sheetnames[0]]
            mylist = []
            cellcount = 0
            checkcount = 0
            for i in range(2,1002):
                if cellcount % 16 == 0:
                    if checkcount != 0:
                        print(sheet.cell(row=i-16, column=1).value, i)
                    checkcount = 0
                cell = sheet.cell(row=i, column=2)
                bgcolor = cell.fill.bgColor.value
                cellcount += 1
                if bgcolor == 64:
                    checkcount += 1
                
            page.update()
        Read = ft.ElevatedButton("Read", on_click = lambda e : read_path(e))
        return ft.View(
            "/page3",
            appbar=ft.AppBar(
                title=ft.Text("Page 3"),
                actions=[
                    ft.IconButton(ft.Icons.HOME, on_click=lambda _: navigate_to("home")),
                    ft.IconButton(ft.Icons.FIBER_NEW, on_click=lambda _: navigate_to("page1")),
                    ft.IconButton(ft.Icons.EDIT_NOTE, on_click=lambda _: navigate_to("page2")),
                    ft.IconButton(ft.Icons.LOOKS_3, on_click=lambda _: navigate_to("page3")),
                    ft.IconButton(ft.Icons.SETTINGS_OUTLINED, on_click=lambda _: navigate_to("setting")),
                ],
            ),
            controls=[
                ft.Text("This is Page 3", size=30),
                ft.Row(
                    controls=[
                        ft.Column(
                            controls=[
                                Path,
                                Read,
                            ],
                            expand=True
                        ),
                    ],
                    scroll=ft.ScrollMode.ALWAYS,
                    expand=True
                ),
            ],
        )

    def setting_view():
        global setting
        optionnum = ft.TextField(
            label="オプションの上限値",
            value=setting[0][0],
        )
        optionlistpath = ft.TextField(
            label="オプションリストのパス",
            value=setting[0][1],
        )
        def save_setting(e):
            optionnum_value = optionnum.value
            optionlistpath_value = optionlistpath.value
            if optionnum_value[0] == "\"" and optionnum_value[-1] == "\"":
                optionnum_value = optionnum_value[1:-1]
            if optionlistpath_value[0] == "\"" and optionlistpath_value[-1] == "\"":
                optionlistpath_value = optionlistpath_value[1:-1]
            cur.execute('UPDATE setting SET optionnumber = ?, optionlistpath = ?', (optionnum_value, optionlistpath_value))
            conn.commit()
            page.update()
        return ft.View(
            "/setting",
            appbar=ft.AppBar(
                title=ft.Text("設定"),
                actions=[
                    ft.IconButton(ft.Icons.HOME, on_click=lambda _: navigate_to("home")),
                    ft.IconButton(ft.Icons.FIBER_NEW, on_click=lambda _: navigate_to("page1")),
                    ft.IconButton(ft.Icons.EDIT_NOTE, on_click=lambda _: navigate_to("page2")),
                    ft.IconButton(ft.Icons.LOOKS_3, on_click=lambda _: navigate_to("page3")),
                    ft.IconButton(ft.Icons.SETTINGS_OUTLINED, on_click=lambda _: navigate_to("setting")),
                ],
            ),
            controls=[
                ft.Row(
                    controls=[
                        ft.Column(
                            controls=[
                                optionnum,
                                optionlistpath,
                                ft.ElevatedButton("Save", on_click=lambda e: save_setting(e)),
                            ],
                            expand=True
                        ),
                    ],
                    scroll=ft.ScrollMode.ALWAYS,
                    expand=True
                ),
            ],
        )

    page.views.append(home_view())
    page.update()

ft.app(target=main)

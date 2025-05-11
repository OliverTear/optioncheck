import flet as ft

import openpyxl
  
  
def main(page: ft.Page):
    page.title = "Multi-Page App"

    def navigate_to(page_name):

        if page_name == "home":
            page.views.clear()
            page.views.append(home_view())
        elif page_name == "make_optioncode":
            page.views.clear()
            page.views.append(make_optioncode_view())
        elif page_name == "edit_optioncode":
            page.views.clear()
            page.views.append(edit_optioncode_view())
        elif page_name == "page3":
            page.views.clear()
            page.views.append(page3_view())
        elif page_name == "page4":
            page.views.clear()
            page.views.append(page4_view())
        page.update()

    def home_view():
        return ft.View(
            "/",
            appbar=ft.AppBar(
                title=ft.Text("Home"),
                actions=[
                    ft.IconButton(ft.Icons.HOME, on_click=lambda _: navigate_to("home")),
                    ft.IconButton(ft.Icons.FIBER_NEW, on_click=lambda _: navigate_to("make_optioncode")),
                    ft.IconButton(ft.Icons.EDIT_NOTE, on_click=lambda _: navigate_to("edit_optioncode")),
                    ft.IconButton(ft.Icons.LOOKS_3, on_click=lambda _: navigate_to("page3")),
                    ft.IconButton(ft.Icons.LOOKS_4, on_click=lambda _: navigate_to("page4")),
                ],
            ),
            controls=[
                ft.Text("Welcome to the Home Page!", size=30),
            ],
        )
    
    def make_optioncode_view():

        
        return ft.View(
            "/make_optioncode",
            appbar=ft.AppBar(
                title=ft.Text("OptionCode.dat Maker"),
                actions=[
                    ft.IconButton(ft.Icons.HOME, on_click=lambda _: navigate_to("home")),
                    ft.IconButton(ft.Icons.FIBER_NEW, on_click=lambda _: navigate_to("make_optioncode")),
                    ft.IconButton(ft.Icons.EDIT_NOTE, on_click=lambda _: navigate_to("edit_optioncode")),
                    ft.IconButton(ft.Icons.LOOKS_3, on_click=lambda _: navigate_to("page3")),
                    ft.IconButton(ft.Icons.LOOKS_4, on_click=lambda _: navigate_to("page4")),
                ],
            ),
            controls=[
                ft.Text("ペーストしたデータを元にOptionCode.datを生成します", size=30),
            ],
        )

    def edit_optioncode_view():
        def on_option(e):
            pass

        def read_data(e):
            global oplist
            oplist = [True]
            opcount = 1
            rows = []
            # for data in alldata:
            #     for i in range(16):
            #         chip = ft.Chip(
            #             label=ft.Text("Option" + str(opcount)),
            #             selected_color= ft.Colors.AMBER,
            #             data = opcount,
            #             selected = (data[i] == "1"),
            #             on_click=lambda e: on_option(e),
            #             width=300,
            #         )
            #         oplist.append(data[i] == "1")
            #         opcount += 1
            #         cells = [ft.DataCell(chip), ft.DataCell(ft.Text(value="Option" + str(i),))]
            #         rows.append(ft.DataRow(cells=cells))
            data_table.rows = rows
            page.update()
        

        header = [ft.DataColumn(ft.Text("ON/OFF")), ft.DataColumn(ft.Text("OptionName"))]
        rows = []
        data_table = ft.DataTable(columns=header, rows=rows)

        editfile_PATH = ft.TextField(
            label="パス",
            value="",
        )
        
        ReadButton = ft.Button("読み込み", on_click = lambda e : read_data(e))
        savebotton = ft.Button("保存", on_click = lambda e : print("保存"))

        return ft.View(
            "/edit_optioncode",
            appbar=ft.AppBar(
                title=ft.Text("Page 2"),
                actions=[
                    ft.IconButton(ft.Icons.HOME, on_click=lambda _: navigate_to("home")),
                    ft.IconButton(ft.Icons.FIBER_NEW, on_click=lambda _: navigate_to("make_optioncode")),
                    ft.IconButton(ft.Icons.EDIT_NOTE, on_click=lambda _: navigate_to("edit_optioncode")),
                    ft.IconButton(ft.Icons.LOOKS_3, on_click=lambda _: navigate_to("page3")),
                    ft.IconButton(ft.Icons.LOOKS_4, on_click=lambda _: navigate_to("page4")),
                ],
            ),
            controls=[
                ft.Text("This is Page 2", size=30),
                ft.Row(
                    controls=[
                        ft.Column(
                            controls=[data_table],
                            scroll=ft.ScrollMode.ALWAYS,
                            expand=True
                        ),
                        ft.Column(
                            controls=[
                                ft.Text("datファイル保存先", size=30),
                                editfile_PATH,
                                ReadButton,
                            ],
                            expand=True
                        ),
                    ],
                    scroll=ft.ScrollMode.ALWAYS,
                    expand=True
                ),
            ],
        )

    def page3_view():
        Path = ft.TextField(
            label="Path",
            value="",
        )
        def read_path(e):
            path = Path.value
            if path[0] == "\"" and path[-1] == "\"":
                path = path[1:-1]
            wb = openpyxl.load_workbook(path)
            sheet = wb[wb.sheetnames[0]]
            mylist = []
            cellcount = 0
            checkcount = 0
            for i in range(2,1002):
                if cellcount % 16 == 0:
                    if checkcount != 0:
                        print(sheet.cell(row=i-16, column=1).value, i)
                    checkcount = 0
                cell = sheet.cell(row=i, column=2)
                bgcolor = cell.fill.bgColor.value
                cellcount += 1
                if bgcolor == 64:
                    checkcount += 1
                
            page.update()
        Read = ft.ElevatedButton("Read", on_click = lambda e : read_path(e))
        return ft.View(
            "/page3",
            appbar=ft.AppBar(
                title=ft.Text("Page 3"),
                actions=[
                    ft.IconButton(ft.Icons.HOME, on_click=lambda _: navigate_to("home")),
                    ft.IconButton(ft.Icons.FIBER_NEW, on_click=lambda _: navigate_to("make_optioncode")),
                    ft.IconButton(ft.Icons.EDIT_NOTE, on_click=lambda _: navigate_to("edit_optioncode")),
                    ft.IconButton(ft.Icons.LOOKS_3, on_click=lambda _: navigate_to("page3")),
                    ft.IconButton(ft.Icons.LOOKS_4, on_click=lambda _: navigate_to("page4")),
                ],
            ),
            controls=[
                ft.Text("This is Page 3", size=30),
                ft.Row(
                    controls=[
                        ft.Column(
                            controls=[
                                Path,
                                Read,
                            ],
                            expand=True
                        ),
                    ],
                    scroll=ft.ScrollMode.ALWAYS,
                    expand=True
                ),
            ],
        )

    def page4_view():
        return ft.View(
            "/page4",
            appbar=ft.AppBar(
                title=ft.Text("Page 4"),
                actions=[
                    ft.IconButton(ft.Icons.HOME, on_click=lambda _: navigate_to("home")),
                    ft.IconButton(ft.Icons.FIBER_NEW, on_click=lambda _: navigate_to("make_optioncode")),
                    ft.IconButton(ft.Icons.EDIT_NOTE, on_click=lambda _: navigate_to("edit_optioncode")),
                    ft.IconButton(ft.Icons.LOOKS_3, on_click=lambda _: navigate_to("page3")),
                    ft.IconButton(ft.Icons.LOOKS_4, on_click=lambda _: navigate_to("page4")),
                ],
            ),
            controls=[
                ft.Text("This is Page 4", size=30),
            ],
        )

    page.views.append(home_view())
    page.update()

ft.app(target=main)
